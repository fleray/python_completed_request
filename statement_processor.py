import json
import logging
from typing import List, Set
import os
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.comments import Comment

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

N1QL_RESERVED_KEYWORDS = ['ADVISE', 'ALL', 'ALTER', 'ANALYZE', 'AND', 'ANY', 'ARRAY', 'AS', 'ASC', 'AT', 'BEGIN', 'BETWEEN', 'BINARY', 'BOOLEAN', 'BREAK', 'BUCKET', 'BUILD', 'BY', 'CACHE', 'CALL', 'CASE', 'CAST', 'CLUSTER', 'COLLATE', 'COLLECTION', 'COMMIT', 'COMMITTED', 'CONNECT', 'CONTINUE', 'CORRELATED', 'COVER', 'CREATE', 'CURRENT', 'CYCLE', 'DATABASE', 'DATASET', 'DATASTORE', 'DECLARE', 'DECREMENT', 'DEFAULT', 'DELETE', 'DERIVED', 'DESC', 'DESCRIBE', 'DISTINCT', 'DO', 'DROP', 'EACH', 'ELEMENT', 'ELSE', 'END', 'ESCAPE', 'EVERY', 'EXCEPT', 'EXCLUDE', 'EXECUTE', 'EXISTS', 'EXPLAIN', 'FALSE', 'FETCH', 'FILTER', 'FIRST', 'FLATTEN', 'FLATTEN_KEYS', 'FLUSH', 'FOLLOWING', 'FOR', 'FORCE', 'FROM', 'FTS', 'FUNCTION', 'GOLANG', 'GRANT', 'GROUP', 'GROUPS', 'GSI', 'HASH', 'HAVING', 'IF', 'IGNORE', 'ILIKE', 'IN', 'INCLUDE', 'INCREMENT', 'INDEX', 'INFER', 'INLINE', 'INNER', 'INSERT', 'INTERSECT', 'INTO', 'IS', 'ISOLATION', 'JAVASCRIPT', 'JOIN', 'KEY', 'KEYS', 'KEYSPACE', 'KNOWN', 'LANGUAGE', 'LAST', 'LATERAL', 'LEFT', 'LET', 'LETTING', 'LEVEL', 'LIKE', 'LIMIT', 'LSM', 'MAP', 'MAPPING', 'MATCHED', 'MATERIALIZED', 'MAXVALUE', 'MERGE', 'MINVALUE', 'MISSING', 'NAMESPACE', 'NEST', 'NEXT', 'NEXTVAL', 'NL', 'NO', 'NOT', 'NTH_VALUE', 'NULL', 'NULLS', 'NUMBER', 'OBJECT', 'OFFSET', 'ON', 'OPTION', 'OPTIONS', 'OR', 'ORDER', 'OTHERS', 'OUTER', 'OVER', 'PARSE', 'PARTITION', 'PASSWORD', 'PATH', 'POOL', 'PRECEDING', 'PREPARE', 'PREV', 'PREVIOUS', 'PREVVAL', 'PRIMARY', 'PRIVATE', 'PRIVILEGE', 'PROBE', 'PROCEDURE', 'PUBLIC', 'RANGE', 'RAW', 'READ', 'REALM', 'RECURSIVE', 'REDUCE', 'RENAME', 'REPLACE', 'RESPECT', 'RESTART', 'RESTRICT', 'RETURN', 'RETURNING', 'REVOKE', 'RIGHT', 'ROLE', 'ROLLBACK', 'ROW', 'ROWS', 'SATISFIES', 'SAVEPOINT', 'SCHEMA', 'SCOPE', 'SELECT', 'SELF', 'SEQUENCE', 'SET', 'SHOW', 'SOME', 'START', 'STATISTICS', 'STRING', 'SYSTEM', 'THEN', 'TIES', 'TO', 'TRAN', 'TRANSACTION', 'TRIGGER', 'TRUE', 'TRUNCATE', 'UNBOUNDED', 'UNDER', 'UNION', 'UNIQUE', 'UNKNOWN', 'UNNEST', 'UNSET', 'UPDATE', 'UPSERT', 'USE', 'USER', 'USERS', 'USING', 'VALIDATE', 'VALUE', 'VALUED', 'VALUES', 'VECTOR', 'VIA', 'VIEW', 'WHEN', 'WHERE', 'WHILE', 'WINDOW', 'WITH', 'WITHIN', 'WORK', 'XOR']

TIME_DEFINITION = {
        'elapsedTime': 'When the request arrives at the server, it is placed into a queue until a worker thread picks it up.\n\nElapsed time is the total time taken for the request, that is the time from when the request was received until the results were returned.\n\nElapsed time includes time spent in the queue, whereas execution time does not.',
        'cpuTime': 'Time spent executing the operator code inside SQL++ query engine, i.e. the actual CPU time consumed by all threads involved.',
        'serviceTime': 'Time spent waiting for another service, such as index or data.\n- For index scan, it is time spent waiting for GSI/indexer.\n- For fetch, it is time spent waiting on the KV store.\n\nA high servTime for a low number of items processed is an indication that the indexer or KV store is stressed.\n\nA high kernTime means there is a downstream issue in the query plan or the query server having many requests to process (so the scheduled waiting time will be more for CPU time).',
}

def handle_in_operator(field: str, value: str, query: str, match_start: int, match_end: int) -> tuple:
    """
    Handle IN operator template creation.
    
    Args:
        field: The field name
        value: The value to process
        query: The original query
        match_start: Start position of the match
        match_end: End position of the match
        
    Returns:
        Tuple of (template_part, new_end_position)
    """
    original_value_str_len = len(str(value))
    
    # For IN operator, we need to find the complete array or list
    if value.startswith('['):
        # Find the complete array including all nested content
        array_match = re.search(r'\[(?:[^\]]*)\]', query[match_start:])
        if array_match:
            value = array_match.group(0)

        # Add the field, operator and placeholder
        template_part = f"{field} IN [?, ?, ...]"
        new_end = match_end + len(str(value)) - original_value_str_len
    elif value.startswith('('):
        # Find the complete list including all nested content
        list_match = re.search(r'\([^)]+\)', query[match_start:])
        if list_match:
            value = list_match.group(0)
    
        # Add the field, operator and placeholder
        template_part = f"{field} IN [?, ?, ...]"
        new_end = match_end + len(str(value)) - original_value_str_len
    else:
        # last case : value represents array, like schedule in : "AND ANY v IN schedule ..."
        # do nothing : ignore
        template_part = f"{field} IN {value}"
        new_end = match_end
    
    return template_part, new_end

def handle_simple_operator(field: str, operator: str, value: str, match_end: int) -> tuple:
    """
    Handle simple operator template creation.
    
    Args:
        field: The field name
        operator: The operator
        value: The value to process
        
    Returns:
        Template part for the simple operator
    """

    new_end = match_end
    original_value = value

    # Remove parenthesis or quotes if present
    if value.startswith("("):
        value = value[1:]
        if value in N1QL_RESERVED_KEYWORDS:
            # Do nothing
            return f"{field} {operator} {original_value}", new_end
        new_end +=  1
    if value.endswith(")"):
        value = value[:-1]
        if value in N1QL_RESERVED_KEYWORDS:
            # Do nothing
            return f"{field} {operator} {original_value}", new_end
        new_end -=  1
    if value.startswith(("'", '"')) and value.endswith(("'", '"')):
        value = value[1:-1]
    
    # Add the field, operator and placeholder
    return f"{field} {operator} ?", new_end

def create_template(query: str) -> str:
    """
    Create a template by replacing values in the query with placeholders.
    
    Args:
        query: The SQL query to template
        
    Returns:
        The templated query
    """
    new_template = ""
    last_end = 0
    
    # Simple value pattern: 'value' or "value" or value
    # For quoted values, can contain spaces: 'my value' or "my value"
    # For unquoted values, no spaces allowed: myvalue
    simple_value = r'(?:[\'"][^\'"]*[\'"]|[^\'",\s]+)'
    # Array pattern: [value1, value2, ...] with optional spaces
    array_value = r'\[(?:[^\]]*)\]'
    
    operator_pattern = f'([a-zA-Z0-9_.]+)\\s*(>=|<=|==|=|>|<| in | like )\\s*({simple_value}|{array_value})'
    
    # Find all matches
    matches = list(re.finditer(operator_pattern, query, re.IGNORECASE))
    
    # Process matches in order
    for match in matches:
        field = match.group(1)
        operator = match.group(2)
        value = match.group(3).strip()
        
        # Add the text before the match
        new_template += query[last_end:match.start()]
        
        if value.startswith("$"): # already "named" or "positional" parametrized value
            new_template += f"{field} {operator} {value}"
            last_end = match.end()
            continue
        
        # Handle different operators
        if operator.lower() == ' in ':
            template_part, last_end = handle_in_operator(field, value, query, match.start(), match.end())
            new_template += template_part
        else:
            template_part, last_end = handle_simple_operator(field, operator, value, match.end())
            new_template += template_part
    
    # Add any remaining text after the last match
    new_template += query[last_end:]
    
    return new_template

def process_positional_args(statement: str, positional_args: List) -> str:
    """
    Process the SQL statement by replacing numbered placeholders ($1, $2, etc.) with values from positional_args.
    """
    def replace_arg(match):
        try:
            index = int(match.group(1)) - 1  # Convert to 0-based index
            if 0 <= index < len(positional_args):
                value = positional_args[index]
                # Wrap string values with single quotes
                if isinstance(value, str):
                    return f"'{value}'"
                return str(value)
            else:
                logging.warning(f"Positional argument index {index + 1} out of range")
                return match.group(0)
        except ValueError:
            logging.error(f"Invalid positional argument format: {match.group(0)}")
            return match.group(0)
    
    # Replace all occurrences of $n with corresponding positional arguments
    pattern = r'\$(\d+)'
    return re.sub(pattern, replace_arg, statement)

def process_named_args(statement: str, named_args: dict) -> str:
    """
    Process the SQL statement by replacing named placeholders ($keyword) with values from named_args.
    Example: 
    - statement: "SELECT * FROM users WHERE name = $name AND age > $age"
    - named_args: {"$name": "John", "$age": 18}
    - result: "SELECT * FROM users WHERE name = 'John' AND age > 18"
    """
    def replace_arg(match):
        try:
            placeholder = match.group(0)  # Get the full match including $ (e.g., "$name")
            if placeholder in named_args:
                value = named_args[placeholder]
                # Wrap string values with single quotes
                if isinstance(value, str):
                    return f"'{value}'"
                return str(value)
            else:
                logging.warning(f"Named argument '{placeholder}' not found in provided arguments")
                return placeholder
        except Exception as e:
            logging.error(f"Error processing named argument: {str(e)}")
            return match.group(0)
    
    # Replace all occurrences of $keyword with corresponding named arguments
    pattern = r'\$\w+'
    return re.sub(pattern, replace_arg, statement)

def process_json_file(file_path: str, use_value_for_parameters: bool) -> List[dict]:
    """
    Read and process the JSON file containing SQL statements and their metadata.
    """
    try:
        with open(file_path, 'r') as f:
            data = json.load(f)
            
        processed_items = []
        
        # Check if data is a list
        if not isinstance(data, list):
            logging.error("Input JSON must be a list of objects")
            return []
            
        for completed_request in data:
            
            # Check if required fields are present
            if 'statement' not in completed_request:
                logging.warning(f"Skipping item missing required statement field: {completed_request}")
                continue
                
            # Process the statement and arguments (positional or named)
            statement = completed_request['statement']
            processed_statement = statement.replace('\n', ' ').replace('<ud>', '').replace('</ud>', '')

            if(use_value_for_parameters):
                positional_args = completed_request.get('positionalArgs', [])
                if(len(positional_args) > 0):
                    processed_statement = process_positional_args(
                        processed_statement, positional_args)
                    
                named_args = completed_request.get('namedArgs', [])
                if(len(named_args) > 0):
                    processed_statement = process_named_args(
                        processed_statement, named_args)
            
            # Create a new item with the processed statement
            processed_item = completed_request.copy()
            processed_item['statement'] = processed_statement

            processed_items.append(processed_item)
            
        return processed_items
        
    except FileNotFoundError:
        logging.error(f"File not found: {file_path}")
        return []
    except json.JSONDecodeError:
        logging.error(f"Invalid JSON format in file: {file_path}")
        return []
    except Exception as e:
        logging.error(f"Error processing file: {str(e)}")
        return []

def convert_to_excel_value(value):
    """
    Convert a value to a format that can be written to Excel.
    """
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, str):
        value = value.replace('µs', 'us')
    return value

def convert_to_seconds(time_str):
    """Convert time string to seconds."""
    if not time_str:
        return 0
    try:
        if isinstance(time_str, (int, float)):
            return float(time_str)
        if 'us' in time_str:
            return float(time_str.replace('us', '')) / 1000000
        if 'µs' in time_str:
            return float(time_str.replace('µs', '')) / 1000000
        if 'ms' in time_str:
            return float(time_str.replace('ms', '')) / 1000
        if 's' in time_str:
            return float(time_str.replace('s', ''))
        if 'm' in time_str:
            return float(time_str.replace('m', '')) * 60
        if 'h' in time_str:
            return float(time_str.replace('h', '')) * 3600
        return float(time_str)
    except (ValueError, TypeError):
        return 0


def convert_to_micro_seconds(time_str):
    """Convert time string to micro seconds."""
    if not time_str:
        return 0
    try:
        if isinstance(time_str, (int, float)):
            return float(time_str)
        if 'us' in time_str:
            return float(time_str.replace('us', ''))
        if 'µs' in time_str:
            return float(time_str.replace('µs', ''))
        if 'ms' in time_str:
            return float(time_str.replace('ms', '')) * 1000
        if 's' in time_str:
            return float(time_str.replace('s', '')) * 100000
        return float(time_str)
    except (ValueError, TypeError):
        return 0

def create_sheet_headers(ws, headers, header_font, header_fill):
    """Create and style headers for a worksheet."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Define comment texts for specific headers
    comment_texts = {
        'elapsedTime': TIME_DEFINITION['elapsedTime'],
        'cpuTime': TIME_DEFINITION['cpuTime'],
        'serviceTime': TIME_DEFINITION['serviceTime'],
        'AVG elapsedTime (s)': TIME_DEFINITION['elapsedTime'],
        'AVG cpuTime (µs)': TIME_DEFINITION['cpuTime'],
        'AVG serviceTime (s)' : TIME_DEFINITION['serviceTime']
    }
    
    # Add comments to specific header cells
    for col_idx, header in enumerate(headers, 1):
        
        if header in comment_texts:
            comment_text = comment_texts[header]
            cell = ws.cell(row=1, column=col_idx)
            # Calculate height based on text length (approximate)
            height = max(50, int(len(comment_text) * 0.6)) # Adjust multiplier as needed
            cell.comment = Comment(comment_text, 'Metric Explanation', width=400, height=height)

def calculate_averages(group):
    """Calculate average values for a group of metrics."""
    return {
        'elapsedTime': sum(group['elapsedTime']) / len(group['elapsedTime']),
        'totalElapsedTime': sum(group['elapsedTime']),
        'cpuTime': sum(group['cpuTime']) / len(group['cpuTime']),
        'serviceTime': sum(group['serviceTime']) / len(group['serviceTime']),
        'resultCount': sum(group['resultCount']) / len(group['resultCount']),
        'resultSize': sum(group['resultSize']) / len(group['resultSize']),
        'count': group['count']
    }

def write_group_data(ws, row_idx, group, averages):
    """Write group data to worksheet."""
    ws.cell(row=row_idx, column=1, value=group['requestTime'])
    ws.cell(row=row_idx, column=2, value=group['statement'])
    ws.cell(row=row_idx, column=3, value=averages['elapsedTime'])
    ws.cell(row=row_idx, column=4, value=averages['totalElapsedTime'])
    ws.cell(row=row_idx, column=5, value=averages['cpuTime'])
    ws.cell(row=row_idx, column=6, value=averages['serviceTime'])
    ws.cell(row=row_idx, column=7, value=averages['resultCount'])
    ws.cell(row=row_idx, column=8, value=averages['resultSize'])
    ws.cell(row=row_idx, column=9, value=averages['count'])

def create_excel_sheets(wb: Workbook, processed_items: List[dict], sheet_title: str, sample_statement: bool = False) -> None:
    """
    Create and populate Excel sheets with processed data.
    
    Args:
        wb: Workbook object
        processed_items: List of processed items
        sheet_title: Title prefix for the sheets (e.g., "Raw" or "Aggregated")
        sample_statement: Whether to add 1 sample statement per template as note for for Normalize Queries Aggregated tab (only).
    """
    # Define headers in the specified order
    headers = [
        'requestTime', 'statement', 'elapsedTime', 'cpuTime', 'serviceTime', 
        'resultCount', 'resultSize', 'phaseCounts', 'phaseOperators', 
        'phaseTimes', 'queryContext', 'remoteAddr', 'requestId', 'errorCount',
        'errors', 'namedArgs', 'n1qlFeatCtrl', 'clientContextID',
        'scanConsistency', 'state', 'statementType', 'useCBO',
        'usedMemory', 'userAgent', 'users', '~qualifier'
    ]
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    
    # Create and setup the "1st sheet" (Raw Results)
    ws_raw = wb.active if sheet_title == "Raw" else wb.create_sheet(title=f"{sheet_title} Queries")
    create_sheet_headers(ws_raw, headers, header_font, header_fill)
    
    # Write data rows
    for row_idx, item in enumerate(processed_items, 2):
        for col_idx, header in enumerate(headers, 1):
            value = convert_to_excel_value(item.get(header, ''))
            ws_raw.cell(row=row_idx, column=col_idx, value=value)
    
    # Create and setup the "2nd sheet" (Aggregated Results)
    ws_agg = wb.create_sheet(title=f"{sheet_title} Queries (Aggregated)")
    
    # Define headers for aggregated sheet
    agg_headers = [
        'requestTime', 'statement TEMPLATE', 'AVG elapsedTime (s)', 
        'TOTAL elapsedTime (s)', 'AVG cpuTime (µs)', 'AVG serviceTime (s)', 
        'AVG resultCount', 'AVG resultSize (bytes)', 'TOTAL count'
    ]
    
    create_sheet_headers(ws_agg, agg_headers, header_font, header_fill)
    
    # Group by statement and calculate averages
    statement_groups = {}
    for item in processed_items:
        statement = item['statement']
        if statement not in statement_groups:
            statement_groups[statement] = {
                'requestTime': item['requestTime'],
                'statement': statement,
                'elapsedTime': [],
                'cpuTime': [],
                'serviceTime': [],
                'resultCount': [],
                'resultSize': [],
                'count': 0
            }
        
        # Add values to the group
        statement_groups[statement]['elapsedTime'].append(convert_to_seconds(item.get('elapsedTime', 0)))
        statement_groups[statement]['cpuTime'].append(convert_to_micro_seconds(item.get('cpuTime', 0)))
        statement_groups[statement]['serviceTime'].append(convert_to_seconds(item.get('serviceTime', 0)))
        statement_groups[statement]['resultCount'].append(float(item.get('resultCount', 0)))
        statement_groups[statement]['resultSize'].append(float(item.get('resultSize', 0)))
        statement_groups[statement]['count'] += 1
    
    # Sort statement_groups by total elapsedTime in descending order
    sorted_groups = sorted(
        statement_groups.items(),
        key=lambda x: sum(x[1]['elapsedTime']),
        reverse=True
    )
    
    # Write aggregated data
    for row_idx, (_, group) in enumerate(sorted_groups, 2):
        averages = calculate_averages(group)
        write_group_data(ws_agg, row_idx, group, averages)
    
    # Add color gradient to TOTAL elapsedTime column
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='FFFF00',
        mid_type='percentile', mid_value=50, mid_color='FFA500',
        end_type='max', end_color='FF0000'
    )
    ws_agg.conditional_formatting.add(
        f'D2:D{ws_agg.max_row}',
        color_scale_rule
    )
    
    # Create and setup the "3rd sheet" (Normalized Queries Aggregated)
    if sheet_title == "Param.":
        ws_normalized = wb.create_sheet(title=f"Normalized Queries (Aggregated)")
        create_sheet_headers(ws_normalized, agg_headers, header_font, header_fill)
        
        # Group by template and calculate averages
        template_groups = {}
        template_to_statements = {}  # Dictionary to store statements for each template
        
        for item in processed_items:
            statement = item['statement']
            template = create_template(statement)
            
            if template not in template_groups:
                template_groups[template] = {
                    'requestTime': item['requestTime'],
                    'statement': template,
                    'elapsedTime': [],
                    'cpuTime': [],
                    'serviceTime': [],
                    'resultCount': [],
                    'resultSize': [],
                    'count': 0
                }
                # Set 1 example statement for this template
                template_to_statements[template] = statement
            
            # Add values to the group
            template_groups[template]['elapsedTime'].append(convert_to_seconds(item.get('elapsedTime', 0)))
            template_groups[template]['cpuTime'].append(convert_to_seconds(item.get('cpuTime', 0)))
            template_groups[template]['serviceTime'].append(convert_to_seconds(item.get('serviceTime', 0)))
            template_groups[template]['resultCount'].append(float(item.get('resultCount', 0)))
            template_groups[template]['resultSize'].append(float(item.get('resultSize', 0)))
            template_groups[template]['count'] += 1
        
        # Sort template_groups by total elapsedTime in descending order
        sorted_templates = sorted(
            template_groups.items(),
            key=lambda x: sum(x[1]['elapsedTime']),
            reverse=True
        )
        
        # Write aggregated data
        for row_idx, (_, group) in enumerate(sorted_templates, 2):
            averages = calculate_averages(group)
            write_group_data(ws_normalized, row_idx, group, averages)
            
            # Add comment only if sample_statement is True
            if sample_statement:
                cell = ws_normalized.cell(row=row_idx, column=2)
                height = max(100, int(len(template_to_statements[group['statement']]) * 0.3)) # Adjust multiplier as needed
                cell.comment = Comment("Example:\n" + template_to_statements[group['statement']], 'Example', height, 600)
        
        # Add color gradient to TOTAL elapsedTime column
        ws_normalized.conditional_formatting.add(
            f'D2:D{ws_normalized.max_row}',
            color_scale_rule
        )
    
    # Adjust column widths for all sheets
    for ws in [ws_raw, ws_agg, ws_normalized if sheet_title == "Param." else None]:
        if ws:
            for col_idx, header in enumerate(ws[1], 1):
                max_length = len(str(header.value))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    if row[0].value:
                        max_length = max(max_length, len(str(row[0].value)))
                ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 100)

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Process computed request statements from a JSON file and generate Excel report to help identify better slow queries.')
    parser.add_argument('input_file', help='Path to the input JSON file (output from computed request)')
    parser.add_argument('--sample-statement', action='store_true', help='In the output Excel file, for Normalize Queries Aggregated tab only, add 1 sample statement per template as an Excel note')
    args = parser.parse_args()
    
    # Process the JSON file
    processed_items = process_json_file(args.input_file, False)
    
    if not processed_items:
        logging.error("No items to process")
        return
        
    # Create output filename
    input_filename = os.path.splitext(os.path.basename(args.input_file))[0]
    output_file = f"output_{input_filename}.xlsx"
    
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default empty sheet
    
    # Create sheets for parametrized results
    create_excel_sheets(wb, processed_items, "Param.", args.sample_statement)
    
    # Process the JSON file with parameter replacement
    processed_items = process_json_file(args.input_file, True)
    
    # Create sheets for valued results
    create_excel_sheets(wb, processed_items, "Valued", args.sample_statement)
    
    # Save the workbook
    wb.save(output_file)
    logging.info(f"Results written to {output_file}")

if __name__ == "__main__":
    main() 